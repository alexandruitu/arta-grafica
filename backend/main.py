"""FastAPI backend for Arta Grafica Production Planning."""
from __future__ import annotations
import os
from typing import Dict, List, Optional
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo

TZ_RO = ZoneInfo("Europe/Bucharest")
from collections import defaultdict
import hashlib as _hashlib
import threading as _threading
from fastapi import FastAPI, Depends, Query, HTTPException, Request, UploadFile, File, Body
from pydantic import BaseModel as _BaseModel
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, StreamingResponse, JSONResponse
import json as _json
from sqlalchemy.orm import Session
from sqlalchemy import func, case as _case

from database import engine, get_db, Base
from models import (
    Comanda, DispatchItem, Operatie, Deficit, Resursa, ProgramResursa,
    PlanificareSesiune, PlanificareRezultat, Setari,
)
from schemas import (
    ComandaOut, DispatchOut, PlanificareOut, GanttTask,
    ResursaOut, ImportResult, PlanningResult, StocArticol, ComandaSummary, FrozenBody, SetareItem,
)
from importer import import_all
from planner import run_planning

Base.metadata.create_all(bind=engine)

def _seed_setari(db_session):
    """Ensure default settings exist. Safe to call concurrently (upsert, no-op on conflict)."""
    from sqlalchemy.dialects.sqlite import insert as _sqlite_insert
    DEFAULTS = {
        "material_threshold": "0.01",
    }
    for cheie, valoare in DEFAULTS.items():
        stmt = _sqlite_insert(Setari).values(cheie=cheie, valoare=valoare)
        stmt = stmt.on_conflict_do_nothing(index_elements=["cheie"])
        db_session.execute(stmt)
    db_session.commit()

from database import SessionLocal as _SessionLocal
with _SessionLocal() as _sess:
    _seed_setari(_sess)

# ── Auth ───────────────────────────────────────────────────────────────────────
_AUTH_USER = os.environ.get("AG_USER", "andrei")
_AUTH_PASS = os.environ.get("AG_PASS", "sarbu1234")
_AUTH_SALT = os.environ.get("AG_SALT", "arta-grafica-2026")
_VALID_TOKEN: str = _hashlib.sha256(
    f"{_AUTH_USER}:{_AUTH_PASS}:{_AUTH_SALT}".encode()
).hexdigest()

_PLAN_LOCK = _threading.Lock()


class PlanOptions(_BaseModel):
    """Options for the planning algorithm."""
    ignore_material: bool = False
    ignore_rank: bool = False

# Status constants — must match planner.py output
PREVIZIONAT_STATUSES = {"previzionat", "previzionat_bt", "previzionat_material", "previzionat_semifabricat"}
PLACED_STATUSES      = {"planned"} | PREVIZIONAT_STATUSES   # ops that got a time slot
BLOCKED_STATUSES     = {"no_bt", "no_material", "no_resource", "blocked_by_rank",
                         "blocat_semifabricat", "blocat_prefabricat"}

app = FastAPI(title="Arta Grafica - Production Planning", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.middleware("http")
async def auth_middleware(request: Request, call_next):
    path = request.url.path
    # Public: login endpoint + all static/SPA assets + CORS preflight
    if request.method == "OPTIONS" or path == "/api/auth/login" or not path.startswith("/api/"):
        return await call_next(request)
    # Protected: all other /api/* routes
    token = request.headers.get("Authorization", "")
    if token == f"Bearer {_VALID_TOKEN}":
        return await call_next(request)
    return JSONResponse({"detail": "Unauthorized"}, status_code=401)


@app.post("/api/auth/login")
def login(body: dict):
    if body.get("username") == _AUTH_USER and body.get("password") == _AUTH_PASS:
        return {"token": _VALID_TOKEN}
    raise HTTPException(status_code=401, detail="Credentiale incorecte")

DATA_DIR = os.path.join(os.path.dirname(__file__), "..", "data")

_EXPECTED_FILES = {
    "Stari comenzi_AS.xlsx",
    "Dispatch List_AS.xlsx",
    "OperatiiWO_AS.xlsx",
    "Lista Deficite_AS.xlsx",
    "Resurse_AS.xlsx",
}


# --- Setari ---
@app.get("/api/setari")
def get_setari(db: Session = Depends(get_db)):
    """Return all settings as {cheie: valoare} dict."""
    rows = db.query(Setari).all()
    return {r.cheie: r.valoare for r in rows}


@app.put("/api/setari")
def update_setari(body: dict = Body(...), db: Session = Depends(get_db)):
    """Upsert settings. Body: {cheie: valoare, ...}"""
    for cheie, valoare in body.items():
        row = db.query(Setari).filter(Setari.cheie == cheie).first()
        if row:
            row.valoare = str(valoare) if valoare is not None else None
        else:
            db.add(Setari(cheie=cheie, valoare=str(valoare) if valoare is not None else None))
    db.commit()
    rows = db.query(Setari).all()
    return {r.cheie: r.valoare for r in rows}


# --- Import ---
@app.post("/api/import", response_model=ImportResult)
async def do_import(
    files: List[UploadFile] = File(...),
    db: Session = Depends(get_db),
):
    received = {f.filename: f for f in files}
    missing = _EXPECTED_FILES - set(received.keys())
    if missing:
        raise HTTPException(status_code=400, detail=f"Fisiere lipsa: {', '.join(sorted(missing))}")
    os.makedirs(DATA_DIR, exist_ok=True)
    for filename, upload in received.items():
        if filename not in _EXPECTED_FILES:
            continue
        content = await upload.read()
        with open(os.path.join(DATA_DIR, filename), "wb") as fp:
            fp.write(content)
    results = import_all(db, DATA_DIR)
    return ImportResult(**results)


# --- Planning ---
@app.post("/api/plan", response_model=PlanningResult)
def do_plan(
    db: Session = Depends(get_db),
    body: PlanOptions = Body(default=PlanOptions()),
):
    """Run planning. Optional JSON body: {ignore_material: bool, ignore_rank: bool}"""
    if not _PLAN_LOCK.acquire(blocking=False):
        raise HTTPException(
            status_code=409,
            detail="O planificare este deja in curs. Asteptati finalizarea ei.",
        )
    try:
        result = run_planning(
            db,
            ignore_material=body.ignore_material,
            ignore_rank=body.ignore_rank,
        )
    finally:
        _PLAN_LOCK.release()
    return PlanningResult(**result)


# --- Comenzi ---
@app.get("/api/comenzi", response_model=List[ComandaOut])
def list_comenzi(
    search: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    stadiu: Optional[str] = Query(None),
    limit: int = Query(100),
    offset: int = Query(0),
    db: Session = Depends(get_db),
):
    q = db.query(Comanda)
    if status:
        q = q.filter(Comanda.status_cda == status)
    if stadiu:
        q = q.filter(Comanda.stadiu_prepress == stadiu)
    if search:
        search_term = f"%{search}%"
        text_filter = (
            (Comanda.client.ilike(search_term)) |
            (Comanda.articol.ilike(search_term)) |
            (Comanda.ref_client.ilike(search_term))
        )
        try:
            cp_int = int(search)
            text_filter = text_filter | (Comanda.cp == cp_int) | (Comanda.cv == cp_int)
        except ValueError:
            pass
        q = q.filter(text_filter)
    return q.order_by(Comanda.data_actualizata_livrare).offset(offset).limit(limit).all()


@app.get("/api/comenzi/{cp}", response_model=ComandaOut)
def get_comanda(cp: int, db: Session = Depends(get_db)):
    c = db.query(Comanda).filter(Comanda.cp == cp).first()
    if not c:
        raise HTTPException(status_code=404, detail="Comanda not found")
    return c


@app.get("/api/comenzi/{cp}/operatii")
def get_comanda_operatii(cp: int, db: Session = Depends(get_db)):
    ops = db.query(DispatchItem).filter(DispatchItem.wo == cp).all()
    sesiune = (
        db.query(PlanificareSesiune)
        .filter(PlanificareSesiune.status == "completed")
        .order_by(PlanificareSesiune.id.desc())
        .first()
    )
    planned_map: dict = {}
    if sesiune:
        for r in db.query(PlanificareRezultat).filter(
            PlanificareRezultat.sesiune_id == sesiune.id,
            PlanificareRezultat.wo == cp,
        ).all():
            planned_map[r.op] = r
    # Pre-load all results in this session for queue analysis
    all_results_in_session: list[PlanificareRezultat] = []
    if sesiune:
        all_results_in_session = db.query(PlanificareRezultat).filter(
            PlanificareRezultat.sesiune_id == sesiune.id,
            PlanificareRezultat.resursa_id.isnot(None),
            PlanificareRezultat.data_start.isnot(None),
            PlanificareRezultat.data_end.isnot(None),
        ).all()

    result = []
    for op in ops:
        r = planned_map.get(op.op)

        # Queue analysis: ops on the same resource scheduled before this one
        coada = []
        if r and r.resursa_id and r.data_start:
            for other in all_results_in_session:
                if (
                    other.resursa_id == r.resursa_id
                    and other.data_end is not None
                    and other.data_end <= r.data_start
                    and other.wo != cp  # exclude same WO
                ):
                    coada.append({
                        "wo": other.wo,
                        "op": other.op,
                        "start": other.data_start.strftime("%Y-%m-%d %H:%M"),
                        "end": other.data_end.strftime("%Y-%m-%d %H:%M"),
                        "durata_ore": round(other.durata_ore, 1),
                    })
            # Sort by start time
            coada.sort(key=lambda x: x["start"])

        # Parse prefabricat info when the op is blocked/previzionat due to a semifabricat
        prefabricat_info = None
        if r and r.status in ("blocat_prefabricat", "blocat_semifabricat", "previzionat_semifabricat") and r.motiv and r.motiv.startswith("prefabricat:"):
            parts = r.motiv.split(":", 2)  # ["prefabricat", "wo1,wo2", "stock_code"]
            try:
                producer_wos = [int(w) for w in parts[1].split(",") if w]
                articol_prefabricat = parts[2] if len(parts) > 2 else None
                producatori = []
                if sesiune:
                    for pwo in producer_wos:
                        prod_result = (
                            db.query(PlanificareRezultat)
                            .filter(
                                PlanificareRezultat.sesiune_id == sesiune.id,
                                PlanificareRezultat.wo == pwo,
                            )
                            .order_by(PlanificareRezultat.data_end.desc().nullslast())
                            .first()
                        )
                        producatori.append({
                            "wo": pwo,
                            "status": prod_result.status if prod_result else "neplanificat",
                            "data_end": prod_result.data_end.strftime("%Y-%m-%d %H:%M")
                                        if prod_result and prod_result.data_end else None,
                        })
                prefabricat_info = {
                    "articol": articol_prefabricat,
                    "producatori": producatori,
                }
            except (IndexError, ValueError):
                pass

        result.append({
            "id": op.id, "cl": op.cl, "wo": op.wo, "op": op.op,
            "descr_op": op.descr_op, "stock_code": op.stock_code,
            "comandat": op.comandat, "q_plan": op.q_plan,
            "p_setup": op.p_setup, "p_runtime": op.p_runtime,
            "r_setup": op.r_setup, "r_runtime": op.r_runtime,
            "q_raportat": op.q_raportat, "q_rest": op.q_rest,
            "data_start_plan": r.data_start.strftime("%Y-%m-%d %H:%M") if r and r.data_start else None,
            "data_end_plan": r.data_end.strftime("%Y-%m-%d %H:%M") if r and r.data_end else None,
            "status_plan": r.status if r else None,
            "resursa_plan": r.resursa_nume if r else None,
            "coada_lungime": len(coada),
            "coada": coada,
            "prefabricat_info": prefabricat_info,
        })
    return result


# --- Dispatch ---
@app.get("/api/dispatch", response_model=List[DispatchOut])
def list_dispatch(
    cl: Optional[str] = Query(None),
    wo: Optional[int] = Query(None),
    limit: int = Query(200),
    offset: int = Query(0),
    db: Session = Depends(get_db),
):
    q = db.query(DispatchItem)
    if cl:
        q = q.filter(DispatchItem.cl == cl)
    if wo:
        q = q.filter(DispatchItem.wo == wo)
    return q.offset(offset).limit(limit).all()


# --- Resurse ---
@app.get("/api/resurse", response_model=List[ResursaOut])
def list_resurse(cl: Optional[str] = Query(None), db: Session = Depends(get_db)):
    q = db.query(Resursa)
    if cl:
        q = q.filter(Resursa.cl == cl)
    return q.all()


@app.get("/api/resurse/centre-lucru")
def list_centre_lucru(db: Session = Depends(get_db)):
    results = db.query(Resursa.cl, Resursa.denumire_cl).distinct().all()
    return [{"cl": r[0], "denumire": r[1]} for r in results]


# --- Planificare results ---
@app.get("/api/planificare/latest")
def get_latest_planning(db: Session = Depends(get_db)):
    sesiune = db.query(PlanificareSesiune).order_by(PlanificareSesiune.id.desc()).first()
    if not sesiune:
        return {"error": "No planning session found. Run POST /api/plan first."}

    results = (
        db.query(PlanificareRezultat)
        .filter(PlanificareRezultat.sesiune_id == sesiune.id)
        .all()
    )

    stats = {}
    for r in results:
        stats[r.status] = stats.get(r.status, 0) + 1

    return {
        "sesiune_id": sesiune.id,
        "created_at": sesiune.created_at.isoformat() if sesiune.created_at else None,
        "status": sesiune.status,
        "total_operatii": sesiune.total_operatii,
        "operatii_planificate": sesiune.operatii_planificate,
        "operatii_neplanificate": sesiune.operatii_neplanificate,
        "breakdown": stats,
    }


@app.get("/api/planificare/gantt", response_model=List[GanttTask])
def get_gantt_data(
    cl: Optional[str] = Query(None),
    wo: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    sesiune = db.query(PlanificareSesiune).order_by(PlanificareSesiune.id.desc()).first()
    if not sesiune:
        return []

    q = (
        db.query(PlanificareRezultat)
        .filter(PlanificareRezultat.sesiune_id == sesiune.id)
        .filter(PlanificareRezultat.status.in_(PLACED_STATUSES))
    )
    if cl:
        q = q.filter(PlanificareRezultat.cl == cl)

    results = q.all()
    if not results:
        return []

    # ── Bulk pre-load to eliminate N+1 ───────────────────────────────────────
    wo_ids = {r.wo for r in results}
    op_ids = {str(r.op) for r in results}

    comanda_map: dict = {
        c.cp: c
        for c in db.query(Comanda).filter(Comanda.cp.in_(wo_ids)).all()
    }

    # WO "contains" filter (applied after comanda_map is built so we can also
    # apply the client/articol search in the same pass)
    if wo:
        results = [r for r in results if wo.lower() in str(r.wo).lower()]
    if search:
        search_lower = search.lower()
        def _matches_search(r):
            if search_lower in str(r.wo).lower():
                return True
            c = comanda_map.get(r.wo)
            if c:
                if c.client and search_lower in c.client.lower():
                    return True
                if c.articol and search_lower in c.articol.lower():
                    return True
            return False
        results = [r for r in results if _matches_search(r)]

    if not results:
        return []

    # Recompute wo_ids / op_ids after optional filtering
    wo_ids = {r.wo for r in results}
    op_ids = {str(r.op) for r in results}

    operatie_map: dict = {
        o.cod: o
        for o in db.query(Operatie).filter(Operatie.cod.in_(op_ids)).all()
    }
    # All placed ops for this session for the WOs in our result set (for dependency computation)
    all_planned = (
        db.query(PlanificareRezultat)
        .filter(PlanificareRezultat.sesiune_id == sesiune.id)
        .filter(PlanificareRezultat.status.in_(PLACED_STATUSES))
        .filter(PlanificareRezultat.wo.in_(wo_ids))
        .all()
    )
    # Index: wo → list[PlanificareRezultat]
    planned_by_wo: dict = {}
    for p in all_planned:
        planned_by_wo.setdefault(p.wo, []).append(p)

    # ── Build Gantt tasks ─────────────────────────────────────────────────────
    tasks = []
    for r in results:
        if not r.data_start or not r.data_end:
            continue
        # Ensure end > start (guard against 0-duration ops causing invisible bars)
        gantt_end = r.data_end
        if gantt_end <= r.data_start:
            gantt_end = r.data_start + timedelta(hours=max(float(r.durata_ore or 1), 0.5))

        comanda = comanda_map.get(r.wo)
        delivery = (comanda.data_actualizata_livrare or comanda.dt_livr_prod) if comanda else None
        is_late = bool(delivery and r.data_end.date() > delivery)

        if r.frozen:
            custom_class = "bar-frozen-late" if is_late else "bar-frozen-ok"
        elif r.status in ("previzionat", "previzionat_bt", "previzionat_material", "previzionat_semifabricat"):
            custom_class = "bar-late" if is_late else "bar-previzionat"
        else:
            custom_class = "bar-late" if is_late else "bar-planned"

        # Dependency computation using pre-loaded data.
        deps = []
        op_catalog = operatie_map.get(str(r.op))
        my_rank = op_catalog.rank if op_catalog else 999

        wo_ops = [p for p in planned_by_wo.get(r.wo, []) if p.op != r.op and p.data_start]

        # Same-resource predecessor: among ops on the same resource that finish
        # at or before this op's start, pick the one that ends latest.
        # This captures sequential chaining on the same machine regardless of rank.
        if r.resursa_id and r.data_start:
            same_res = [
                p for p in wo_ops
                if p.resursa_id == r.resursa_id and p.data_end and p.data_end <= r.data_start
            ]
            if same_res:
                direct_pred = max(same_res, key=lambda p: p.data_end)
                deps.append(f"{direct_pred.wo}-{direct_pred.op}")
            else:
                # Fall back to rank-based dependency
                if my_rank > 1:
                    for p in wo_ops:
                        p_cat = operatie_map.get(str(p.op))
                        if p_cat and p_cat.rank < my_rank:
                            deps.append(f"{p.wo}-{p.op}")
        else:
            if my_rank > 1:
                for p in wo_ops:
                    p_cat = operatie_map.get(str(p.op))
                    if p_cat and p_cat.rank < my_rank:
                        deps.append(f"{p.wo}-{p.op}")

        tasks.append(GanttTask(
            id=f"{r.wo}-{r.op}",
            name=f"WO:{r.wo} OP:{r.op} ({r.cl})",
            start=r.data_start.strftime("%Y-%m-%d %H:%M"),
            end=gantt_end.strftime("%Y-%m-%d %H:%M"),
            progress=0,
            dependencies=",".join(deps),
            custom_class=custom_class,
            wo=r.wo,
            op=r.op,
            cl=r.cl,
            resursa=r.resursa_nume,
            status=r.status,
            client=comanda.client if comanda else None,
            articol=comanda.articol if comanda else None,
        ))

    # Sort by WO then by planned start date so rank order is visible in Gantt rows
    tasks.sort(key=lambda t: (t.wo, t.start))
    return tasks


# ── Excel export ──────────────────────────────────────────────────────────────
@app.get("/api/planificare/export-xlsx")
def export_planning_xlsx(
    status: Optional[str] = Query(None),
    cl: Optional[str] = Query(None),
    resursa: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    """Export current planning results as Excel (.xlsx).
    Optional query filters: status, cl, resursa, search (client/articol/WO).
    """
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    sesiune = db.query(PlanificareSesiune).order_by(PlanificareSesiune.id.desc()).first()
    if not sesiune:
        raise HTTPException(status_code=404, detail="Nicio sesiune de planificare gasita.")

    q = (
        db.query(PlanificareRezultat)
        .filter(PlanificareRezultat.sesiune_id == sesiune.id)
    )
    if status:
        statuses = [s.strip() for s in status.split(",")]
        q = q.filter(PlanificareRezultat.status.in_(statuses))
    if cl:
        q = q.filter(PlanificareRezultat.cl == cl)
    if resursa:
        q = q.filter(PlanificareRezultat.resursa_nume == resursa)

    results = q.order_by(PlanificareRezultat.wo, PlanificareRezultat.data_start).all()

    # Bulk-load comenzi for client/articol/delivery info
    wo_ids = {r.wo for r in results}
    comanda_map = {c.cp: c for c in db.query(Comanda).filter(Comanda.cp.in_(wo_ids)).all()}

    # Optional search filter (applied in-memory after comanda_map is ready)
    if search:
        sl = search.lower()
        def _match(r):
            if sl in str(r.wo).lower():
                return True
            c = comanda_map.get(r.wo)
            if c:
                if c.client and sl in c.client.lower():
                    return True
                if c.articol and sl in c.articol.lower():
                    return True
            return False
        results = [r for r in results if _match(r)]

    STATUS_LABEL_RO = {
        "planned":                  "Planificat",
        "previzionat_bt":           "Previzionat (BT lipsă)",
        "previzionat_material":     "Previzionat (material insuficient)",
        "previzionat_semifabricat": "Previzionat (semifabricat în producție)",
        "no_material":              "Fără Material",
        "no_bt":                    "Fără BT",
        "blocked_by_rank":          "Blocat Rank",
        "no_resource":              "Fără Resursă",
        "blocat_semifabricat":      "Blocat – semifabricat neplanificat",
        "blocat_prefabricat":       "Blocat – prefabricat",
        "completed":                "Finalizat",
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "Planificare"

    HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
    ALT_FILL    = PatternFill("solid", fgColor="EFF4FB")
    BORDER_SIDE = Side(style="thin", color="CCCCCC")
    CELL_BORDER = Border(left=BORDER_SIDE, right=BORDER_SIDE, top=BORDER_SIDE, bottom=BORDER_SIDE)

    STATUS_ROW_COLOR = {
        "planned":                  "D4EDDA",
        "previzionat_bt":           "CCE5FF",
        "previzionat_material":     "E2D9F3",
        "previzionat_semifabricat": "D6EAF8",
        "no_material":              "F8D7DA",
        "no_bt":                    "FFF3CD",
        "blocked_by_rank":          "FFE5CC",
        "no_resource":              "F8D7DA",
        "blocat_semifabricat":      "F8D7DA",
        "blocat_prefabricat":       "F8D7DA",
        "completed":                "E9ECEF",
    }

    headers = [
        "WO", "Operație", "CL", "Resursă",
        "Client", "Articol",
        "Start", "Sfârșit", "Durată (h)",
        "Status", "Motiv",
        "Data livrare", "Întârziat", "Înghețat",
    ]
    col_widths = [9, 9, 18, 20, 24, 28, 18, 18, 11, 28, 36, 14, 10, 9]

    ws.row_dimensions[1].height = 18
    for col_idx, (hdr, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=hdr)
        cell.font   = HEADER_FONT
        cell.fill   = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        cell.border = CELL_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"

    for row_idx, r in enumerate(results, start=2):
        comanda = comanda_map.get(r.wo)
        delivery = (comanda.data_actualizata_livrare or comanda.dt_livr_prod) if comanda else None
        if delivery and r.data_end:
            is_late = r.data_end.date() > delivery
        else:
            is_late = False

        row_fill_hex = STATUS_ROW_COLOR.get(r.status, "FFFFFF")
        row_fill = PatternFill("solid", fgColor=row_fill_hex) if row_idx % 2 == 0 else (
            PatternFill("solid", fgColor=row_fill_hex)
        )
        # Alternate row shading: use slightly lighter shade on odd rows when status matches
        # For simplicity just use status color for all rows

        row_data = [
            r.wo,
            str(r.op),
            r.cl or "",
            r.resursa_nume or "",
            comanda.client if comanda else "",
            comanda.articol if comanda else "",
            r.data_start.strftime("%Y-%m-%d %H:%M") if r.data_start else "",
            r.data_end.strftime("%Y-%m-%d %H:%M") if r.data_end else "",
            round(r.durata_ore, 2) if r.durata_ore else 0,
            STATUS_LABEL_RO.get(r.status, r.status),
            r.motiv or "",
            delivery.strftime("%Y-%m-%d") if delivery else "",
            "DA" if is_late else "NU",
            "DA" if r.frozen else "NU",
        ]
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill   = row_fill
            cell.border = CELL_BORDER
            cell.alignment = Alignment(vertical="center")
            if col_idx in (7, 8):  # date columns
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if col_idx in (9,):    # numeric
                cell.alignment = Alignment(horizontal="right", vertical="center")

    # Auto-filter on header row
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from urllib.parse import quote as _quote
    filename = f"planificare_{sesiune.created_at.strftime('%Y%m%d_%H%M') if sesiune.created_at else 'export'}.xlsx"
    encoded_name = _quote(filename)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_name}",
        },
    )


@app.get("/api/planificare/board")
def get_board_data(db: Session = Depends(get_db)):
    """Resource-centric board view with nested CL groups and 5-color status coding.
    Shows all ops that have a scheduled slot (planned, previzionat/Fara BT, frozen)."""
    sesiune = db.query(PlanificareSesiune).order_by(PlanificareSesiune.id.desc()).first()
    if not sesiune:
        return {"groups": [], "items": []}

    # Include all ops with a scheduled slot — not just "planned"
    results = (
        db.query(PlanificareRezultat)
        .filter(PlanificareRezultat.sesiune_id == sesiune.id)
        .filter(PlanificareRezultat.resursa_id.isnot(None))
        .filter(PlanificareRezultat.data_start.isnot(None))
        .all()
    )

    # Cache comenzi
    comanda_cache: dict = {}
    for r in results:
        if r.wo not in comanda_cache:
            comanda_cache[r.wo] = db.query(Comanda).filter(Comanda.cp == r.wo).first()

    # Resources used in this session
    resurse_ids = {r.resursa_id for r in results}
    resurse_map = {r.id: r for r in db.query(Resursa).filter(Resursa.id.in_(resurse_ids)).all()}

    # Build nested groups: CL parent → resource children
    cl_to_res: dict = defaultdict(list)
    sorted_resurse = sorted(resurse_map.values(), key=lambda x: (x.cl or "", x.resursa or ""))
    for r in sorted_resurse:
        cl_to_res[r.cl or "?"].append(str(r.id))

    groups: list = []
    for cl, res_ids in sorted(cl_to_res.items()):
        groups.append({
            "id": f"cl__{cl}",
            "content": cl,
            "nestedGroups": res_ids,
            "cl": cl,
            "isParent": True,
        })
    for r in sorted_resurse:
        groups.append({
            "id": str(r.id),
            "content": r.resursa or str(r.id),
            "cl": r.cl or "",
            "isParent": False,
        })

    # 5-color logic — status values match planner.py output exactly
    def item_color(op: PlanificareRezultat, is_late: bool) -> str:
        if op.frozen:
            # portocaliu = frozen dar imposibil (fara material sau blocat de rank)
            return "#ea580c" if op.status in ("no_material", "blocked_by_rank") else "#7c3aed"
        if op.status in ("previzionat", "previzionat_bt", "previzionat_material", "previzionat_semifabricat"):
            return "#2563eb"   # albastru — previzionat (fara BT / material / semifabricat)
        if is_late:
            return "#dc2626"   # rosu — intarziat
        return "#16a34a"       # verde — planificat, neintarziat

    status_label_map = {
        "planned":                    "Planificat",
        "previzionat":                "Previzionat",
        "previzionat_bt":             "Previzionat (fără BT)",
        "previzionat_material":       "Previzionat (fără material)",
        "previzionat_semifabricat":   "Previzionat (semifabricat în producție)",
        "no_bt":                      "Blocat – fără BT",
        "no_material":                "Blocat – fără material",
        "no_resource":                "Blocat – fără resursă",
        "blocked_by_rank":            "Blocat – rank",
        "blocat_semifabricat":        "Blocat – semifabricat neplanificat",
        "blocat_prefabricat":         "Blocat – prefabricat",
    }

    items: list = []
    for op in results:
        comanda = comanda_cache.get(op.wo)
        delivery = None
        if comanda:
            delivery = comanda.data_actualizata_livrare or comanda.dt_livr_prod

        # Use planner's stored times directly (Varianta B — hour-level precision)
        start_dt = op.data_start
        end_dt = op.data_end
        if not end_dt or (end_dt - start_dt).total_seconds() < 900:
            end_dt = start_dt + timedelta(hours=max(op.durata_ore, 0.25))

        is_late = bool(delivery and end_dt.date() > delivery)
        color = item_color(op, is_late)
        slabel = status_label_map.get(op.status or "", op.status or "?")

        tooltip = (
            f"<b>WO: {op.wo}</b><br>"
            f"OP: {op.op} &nbsp;|&nbsp; CL: {op.cl}<br>"
            f"Resursă: {op.resursa_nume or '-'}<br>"
            f"Status: {slabel}" + (" 🔒" if op.frozen else "") + "<br>"
            f"Durată: {op.durata_ore:.1f} h<br>"
            f"Client: {comanda.client if comanda else '-'}<br>"
            f"Articol: {(comanda.articol or '')[:60] if comanda else '-'}<br>"
            f"Livrare: {delivery or '-'}"
            + ("<br><span style='color:#dc2626;font-weight:bold'>⚠ ÎNTÂRZIAT</span>" if is_late else "")
        )

        items.append({
            "id": op.id,                              # PlanificareRezultat.id — unique
            "group": str(op.resursa_id),
            "start": start_dt.strftime("%Y-%m-%dT%H:%M:%S"),
            "end": end_dt.strftime("%Y-%m-%dT%H:%M:%S"),
            "content": f"WO:{op.wo} OP:{op.op}",
            "title": tooltip,
            "style": f"background-color:{color};border-color:{color};color:#fff;",
            # Extra fields for frontend filtering & sidebar
            "result_id": op.id,
            "wo": op.wo,
            "op": op.op,
            "cl": op.cl or "",
            "durata_ore": op.durata_ore,
            "late": is_late,
            "frozen": op.frozen,
            "status": op.status or "",
            "client": comanda.client if comanda else "",
            "articol": comanda.articol if comanda else "",
        })

    return {"groups": groups, "items": items}


@app.get("/api/planificare/operatii")
def get_planning_results(
    cl: Optional[str] = Query(None),
    wo: Optional[int] = Query(None),
    status: Optional[str] = Query(None),
    limit: int = Query(5000),   # high default — client-side filtering handles the rest
    offset: int = Query(0),
    db: Session = Depends(get_db),
):
    sesiune = db.query(PlanificareSesiune).order_by(PlanificareSesiune.id.desc()).first()
    if not sesiune:
        return []

    q = db.query(PlanificareRezultat).filter(PlanificareRezultat.sesiune_id == sesiune.id)
    if cl:
        q = q.filter(PlanificareRezultat.cl == cl)
    if wo:
        q = q.filter(PlanificareRezultat.wo == wo)
    if status:
        # "previzionat" as a filter value matches all previzionat sub-types
        if status == "previzionat":
            q = q.filter(PlanificareRezultat.status.in_(list(PREVIZIONAT_STATUSES)))
        else:
            q = q.filter(PlanificareRezultat.status == status)

    rows = q.order_by(PlanificareRezultat.data_start.asc().nullslast()).offset(offset).limit(limit).all()

    # Enrich with client + articol from Comanda (batch lookup)
    wo_ids = {r.wo for r in rows}
    comanda_map: dict[int, Comanda] = {
        c.cp: c for c in db.query(Comanda).filter(Comanda.cp.in_(wo_ids)).all()
    }

    result = []
    for r in rows:
        c = comanda_map.get(r.wo)
        result.append({
            "id":           r.id,
            "sesiune_id":   r.sesiune_id,
            "dispatch_id":  r.dispatch_id,
            "wo":           r.wo,
            "op":           r.op,
            "cl":           r.cl,
            "resursa_id":   r.resursa_id,
            "resursa_nume": r.resursa_nume,
            "data_start":   r.data_start.isoformat() if r.data_start else None,
            "data_end":     r.data_end.isoformat()   if r.data_end   else None,
            "durata_ore":   r.durata_ore,
            "frozen":       r.frozen,
            "status":       r.status,
            "motiv":        r.motiv,
            # enriched
            "client":       c.client  if c else None,
            "articol":      c.articol if c else None,
        })
    return result


@app.get("/api/planificare/stats")
def get_planning_stats(db: Session = Depends(get_db)):
    """Returns per-status counts for the latest planning session."""
    sesiune = db.query(PlanificareSesiune).order_by(PlanificareSesiune.id.desc()).first()
    if not sesiune:
        return {"total": 0, "planned": 0, "previzionat": 0, "no_material": 0,
                "no_bt": 0, "blocked_by_rank": 0, "no_resource": 0}
    rows = (
        db.query(PlanificareRezultat.status, func.count(PlanificareRezultat.id))
        .filter(PlanificareRezultat.sesiune_id == sesiune.id)
        .group_by(PlanificareRezultat.status)
        .all()
    )
    counts = {status: cnt for status, cnt in rows}
    total = sum(counts.values())
    # Sum all previzionat sub-types together
    previzionat_total = sum(counts.get(s, 0) for s in PREVIZIONAT_STATUSES)
    return {
        "total": total,
        "planned": counts.get("planned", 0),
        "previzionat": previzionat_total,
        "no_material": counts.get("no_material", 0),
        "no_bt": counts.get("no_bt", 0),
        "blocked_by_rank": counts.get("blocked_by_rank", 0),
        "no_resource": counts.get("no_resource", 0),
    }


@app.get("/api/planificare/by-comanda", response_model=Dict[str, ComandaSummary])
def get_planning_by_comanda(db: Session = Depends(get_db)):
    """Per-WO planning summary keyed by WO string.
    Note: status_material uses global (non-sequential) stock view — useful approximation."""
    sesiune = db.query(PlanificareSesiune).order_by(PlanificareSesiune.id.desc()).first()
    if not sesiune:
        return {}

    all_results = (
        db.query(PlanificareRezultat)
        .filter(PlanificareRezultat.sesiune_id == sesiune.id)
        .all()
    )
    by_wo: dict = {}
    for r in all_results:
        by_wo.setdefault(r.wo, []).append(r)

    wo_ids = set(by_wo.keys())
    comanda_map = {
        c.cp: c
        for c in db.query(Comanda).filter(Comanda.cp.in_(wo_ids)).all()
    }

    # Global stock per article
    stoc_q = db.query(
        Deficit.articol,
        func.max(Deficit.sold_actual).label("sold"),
        func.sum(_case((Deficit.tip_rezervare == "B", Deficit.cantitate), else_=0)).label("rez"),
        func.sum(_case((Deficit.tip_rezervare == "A", Deficit.cantitate), else_=0)).label("aprov"),
    ).group_by(Deficit.articol).all()
    global_stoc = {
        r[0]: {
            "disponibil": (r[1] or 0) + (r[2] or 0),
            "disponibil_final": (r[1] or 0) + (r[2] or 0) + (r[3] or 0),
        }
        for r in stoc_q
    }

    # Articles needed per WO (B-type deficit records)
    wo_arts_q = db.query(Deficit.pe_comanda, Deficit.articol).filter(
        Deficit.tip_rezervare == "B"
    ).all()
    wo_articles: dict = {}
    for pe_comanda, art in wo_arts_q:
        if pe_comanda:
            wo_articles.setdefault(str(pe_comanda), set()).add(art)

    summaries: dict = {}
    for wo, ops in by_wo.items():
        planned_ends = [
            r.data_end for r in ops
            if r.status in PLACED_STATUSES and r.data_end
        ]
        data_planificare = max(planned_ends).date() if planned_ends else None

        comanda = comanda_map.get(wo)
        data_livrare = (comanda.data_actualizata_livrare or comanda.dt_livr_prod) if comanda else None
        intarziere_zile: Optional[int] = None
        if data_planificare and data_livrare:
            intarziere_zile = (data_planificare - data_livrare).days

        statuses = {r.status for r in ops}
        has_previzionat = bool(statuses & PREVIZIONAT_STATUSES)
        has_placed      = bool(statuses & PLACED_STATUSES)
        if statuses <= {"planned"}:
            status_planificare = "Planificat"
        elif has_previzionat and statuses <= PLACED_STATUSES:
            status_planificare = "Previzionat"
        elif statuses <= BLOCKED_STATUSES:
            status_planificare = "Blocat"
        elif has_placed:
            status_planificare = "Partial"
        else:
            status_planificare = "Blocat"

        articles_needed = wo_articles.get(str(wo), set())
        # Prioritate: daca planificatorul a blocat WO-ul din lipsa de material → Lipsa
        if any(r.status == "no_material" for r in ops):
            status_material = "Lipsa"
        else:
            status_material = "Disponibil"
            for art in articles_needed:
                s = global_stoc.get(art, {"disponibil": 0.0, "disponibil_final": 0.0})
                if s["disponibil_final"] < 0:
                    status_material = "Lipsa"
                    break
                elif s["disponibil"] < 0 and status_material == "Disponibil":
                    status_material = "In aprovizionare"

        summaries[str(wo)] = {
            "data_planificare": data_planificare.isoformat() if data_planificare else None,
            "intarziere_zile": intarziere_zile,
            "status_planificare": status_planificare,
            "status_material": status_material,
        }

    return summaries


@app.patch("/api/planificare/operatii/{result_id}/start")
def set_start(result_id: int, body: dict, db: Session = Depends(get_db)):
    """Set manual start time and auto-freeze the operation. Keeps duration unchanged.
    Works on any operation — blocked ones are promoted to 'planned' and frozen."""
    from datetime import timedelta as _td
    r = db.query(PlanificareRezultat).filter(PlanificareRezultat.id == result_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="Rezultat planificare negasit")
    data_start_str = body.get("data_start", "")
    try:
        new_start = datetime.strptime(data_start_str, "%Y-%m-%d %H:%M").replace(tzinfo=TZ_RO)
    except ValueError:
        raise HTTPException(status_code=400, detail="Format invalid. Foloseste: YYYY-MM-DD HH:MM")
    r.data_start = new_start
    durata = float(r.durata_ore) if r.durata_ore and r.durata_ore > 0 else 1.0
    r.data_end = new_start + _td(hours=durata)
    r.frozen = True
    # If the operation was blocked, promote it to planned so it shows on Gantt
    if r.status not in PLACED_STATUSES:
        r.status = "planned"
        r.motiv = f"Planificat manual (anterior: {r.status})"
    db.commit()
    return {
        "id": r.id, "frozen": r.frozen, "status": r.status,
        "data_start": r.data_start.strftime("%Y-%m-%d %H:%M"),
        "data_end": r.data_end.strftime("%Y-%m-%d %H:%M"),
    }


@app.get("/api/planificare/frozen")
def get_frozen(db: Session = Depends(get_db)):
    """List all frozen operations from the latest completed session."""
    sesiune = (
        db.query(PlanificareSesiune)
        .filter(PlanificareSesiune.status == "completed")
        .order_by(PlanificareSesiune.id.desc())
        .first()
    )
    if not sesiune:
        return []
    frozen = db.query(PlanificareRezultat).filter(
        PlanificareRezultat.sesiune_id == sesiune.id,
        PlanificareRezultat.frozen == True,
    ).order_by(PlanificareRezultat.data_start).all()
    wo_ids = {r.wo for r in frozen}
    cm = {c.cp: c for c in db.query(Comanda).filter(Comanda.cp.in_(wo_ids)).all()} if wo_ids else {}
    return [
        {
            "id": r.id, "wo": r.wo, "op": r.op, "cl": r.cl,
            "resursa_nume": r.resursa_nume, "durata_ore": r.durata_ore,
            "data_start": r.data_start.strftime("%Y-%m-%d %H:%M") if r.data_start else None,
            "data_end": r.data_end.strftime("%Y-%m-%d %H:%M") if r.data_end else None,
            "client": cm.get(r.wo).client if cm.get(r.wo) else None,
            "articol": cm.get(r.wo).articol if cm.get(r.wo) else None,
        }
        for r in frozen
    ]


@app.patch("/api/planificare/operatii/{result_id}/frozen")
def set_frozen(result_id: int, body: FrozenBody, db: Session = Depends(get_db)):
    """Freeze or unfreeze a planned operation. Frozen ops survive replanning."""
    r = db.query(PlanificareRezultat).filter(PlanificareRezultat.id == result_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="Rezultat planificare negasit")
    if r.status not in PLACED_STATUSES:
        raise HTTPException(
            status_code=400,
            detail=f"Doar operatiile planificate pot fi frozen (status curent: {r.status})",
        )
    r.frozen = body.frozen
    db.commit()
    return {"id": r.id, "frozen": r.frozen, "status": r.status}


# --- Stoc ---
@app.get("/api/stoc", response_model=List[StocArticol])
def get_stoc(
    search: Optional[str] = Query(None),
    limit: int = Query(5000),
    db: Session = Depends(get_db),
):
    # B-type = reservations (negative cantitate)
    # A-type = incoming stock / aprovizionare (positive cantitate)
    q = db.query(
        Deficit.articol,
        func.max(Deficit.sold_actual).label("sold_actual"),
        func.sum(
            _case((Deficit.tip_rezervare == "B", Deficit.cantitate), else_=0)
        ).label("total_rezervat"),
        func.sum(
            _case((Deficit.tip_rezervare == "A", Deficit.cantitate), else_=0)
        ).label("total_aprovizionare"),
    ).group_by(Deficit.articol)

    if search:
        q = q.filter(Deficit.articol.ilike(f"%{search}%"))

    results = q.order_by(Deficit.articol).limit(limit).all()
    return [
        StocArticol(
            articol=r[0],
            sold_actual=r[1] or 0,
            # B-type cantitate values are negative in DB; store as positive magnitude
            total_rezervat=abs(r[2] or 0),
            total_aprovizionare=r[3] or 0,
            disponibil=(r[1] or 0) - abs(r[2] or 0),
            disponibil_final=(r[1] or 0) - abs(r[2] or 0) + (r[3] or 0),
        )
        for r in results
    ]


@app.get("/api/stats")
def get_stats(db: Session = Depends(get_db)):
    total_comenzi = db.query(Comanda).count()
    comenzi_active = db.query(Comanda).filter(Comanda.status_cda == "LIBER").count()
    comenzi_stop = db.query(Comanda).filter(Comanda.status_cda == "STOP").count()
    total_dispatch = db.query(DispatchItem).count()
    total_resurse = db.query(Resursa).count()

    stadiu_counts = (
        db.query(Comanda.stadiu_prepress, func.count(Comanda.id))
        .group_by(Comanda.stadiu_prepress)
        .all()
    )

    # Planning-based counts from latest completed session
    comenzi_intarziate = 0
    comenzi_blocate = 0
    sesiune = (
        db.query(PlanificareSesiune)
        .filter(PlanificareSesiune.status == "completed")
        .order_by(PlanificareSesiune.id.desc())
        .first()
    )
    if sesiune:
        wo_delivery = {
            c.cp: (c.data_actualizata_livrare or c.dt_livr_prod)
            for c in db.query(Comanda).all()
        }
        BLOCKED_STATUSES = {"no_material", "no_resource", "no_bt", "blocked_by_rank",
                             "blocat_semifabricat", "blocat_prefabricat"}
        by_wo: dict = {}
        for r in db.query(PlanificareRezultat).filter(
            PlanificareRezultat.sesiune_id == sesiune.id
        ).all():
            by_wo.setdefault(r.wo, []).append(r)
        for wo, ops in by_wo.items():
            if any(r.status in BLOCKED_STATUSES for r in ops):
                comenzi_blocate += 1
            planned_ends = [r.data_end for r in ops if r.status in PLACED_STATUSES and r.data_end]
            if planned_ends:
                max_end = max(planned_ends).date()
                delivery = wo_delivery.get(wo)
                if delivery and max_end > delivery:
                    comenzi_intarziate += 1

    return {
        "total_comenzi": total_comenzi,
        "comenzi_active": comenzi_active,
        "comenzi_stop": comenzi_stop,
        "comenzi_intarziate": comenzi_intarziate,
        "comenzi_blocate": comenzi_blocate,
        "total_dispatch": total_dispatch,
        "total_resurse": total_resurse,
        "stadiu_prepress": {s[0]: s[1] for s in stadiu_counts},
    }


# ── AI Assistant ──────────────────────────────────────────────────────────────
try:
    import anthropic as _anthropic
    _ANTHROPIC_OK = True
except ImportError:
    _ANTHROPIC_OK = False

_AI_SYSTEM_PROMPT = """Ești un asistent specializat EXCLUSIV în planificarea producției pentru o tipografie.

REGULI STRICTE — respectă-le întotdeauna:
1. Analizezi DOAR datele furnizate în secțiunea CONTEXT. Nu inventa, nu extrapola, nu presupune.
2. Dacă datele sunt insuficiente pentru o concluzie, spune explicit: "Date insuficiente pentru această concluzie."
3. Răspunsuri în română, structurate, max 300 cuvinte.
4. Folosește bullet points (•) pentru liste.
5. Numerele din răspuns trebuie să corespundă EXACT cu datele din context — niciun număr inventat.
6. Nu oferi sfaturi generice de management — doar analiză specifică pe datele prezente.
7. Dacă o problemă are multiple cauze posibile din date, menționează-le pe TOATE."""


def _build_ai_context(question_id: str, db: Session) -> tuple[str, str]:
    """Returns (question_text, context_text) for the AI endpoint."""
    from datetime import date
    today = date.today()

    # ── Shared: stats + latest planning session ───────────────────────────────
    total_comenzi = db.query(Comanda).count()
    comenzi_active = db.query(Comanda).filter(Comanda.status_cda == "LIBER").count()
    comenzi_stop   = db.query(Comanda).filter(Comanda.status_cda == "STOP").count()

    sesiune = db.query(PlanificareSesiune).order_by(PlanificareSesiune.id.desc()).first()
    plan_stats: dict = {}
    if sesiune:
        for r in db.query(PlanificareRezultat).filter(
            PlanificareRezultat.sesiune_id == sesiune.id
        ).all():
            plan_stats[r.status] = plan_stats.get(r.status, 0) + 1

    plan_line = (
        f"Planificate:{plan_stats.get('planned',0)} | "
        f"FaraMaterial:{plan_stats.get('no_material',0)} | "
        f"BlocateRank:{plan_stats.get('blocked_by_rank',0)} | "
        f"FaraResursa:{plan_stats.get('no_resource',0)} | "
        f"FaraBT:{plan_stats.get('no_bt',0)}"
    ) if sesiune else "Nicio sesiune de planificare disponibila"
    ses_label = (
        f"Sesiunea #{sesiune.id} din "
        f"{sesiune.created_at.strftime('%d.%m.%Y %H:%M')}"
    ) if sesiune else "N/A"

    # ── DASHBOARD: Rezumă situația ────────────────────────────────────────────
    if question_id == "dashboard_summary":
        question = "Rezumă situația curentă de producție și principalele puncte de atenție."
        stadiu_counts = (
            db.query(Comanda.stadiu_prepress, func.count(Comanda.id))
            .group_by(Comanda.stadiu_prepress)
            .all()
        )
        stadiu_str = " | ".join(
            f"{s or 'N/A'}:{c}" for s, c in stadiu_counts
        )
        context = (
            f"DATA: {today}\n"
            f"PLANIFICARE: {ses_label}\n\n"
            f"COMENZI:\n"
            f"- Total: {total_comenzi}\n"
            f"- Active (LIBER): {comenzi_active}\n"
            f"- STOP: {comenzi_stop}\n\n"
            f"REZULTATE PLANIFICARE: {plan_line}\n\n"
            f"STADIU PREPRESS: {stadiu_str}"
        )
        return question, context

    # ── DASHBOARD: Comenzi urgente ────────────────────────────────────────────
    elif question_id == "dashboard_urgent":
        question = "Care sunt comenzile urgente sau cu termenul de livrare depășit?"
        comenzi = (
            db.query(Comanda)
            .filter(Comanda.status_cda == "LIBER")
            .order_by(Comanda.data_actualizata_livrare)
            .limit(40)
            .all()
        )
        overdue, next7, next30 = [], [], []
        for c in comenzi:
            d = c.data_actualizata_livrare or c.dt_livr_prod
            if not d:
                continue
            delta = (d - today).days
            line = (
                f"WO:{c.cp} | {(c.client or 'N/A')[:22]} | "
                f"{(c.articol or '')[:32]} | Livr:{d}"
            )
            if delta < 0:
                overdue.append(f"{line} | DEPASIT {abs(delta)}z")
            elif delta <= 7:
                next7.append(f"{line} | {delta}z")
            elif delta <= 30:
                next30.append(f"{line} | {delta}z")

        context = (
            f"DATA: {today}\n"
            f"COMENZI ACTIVE: {comenzi_active}\n\n"
            f"DEPASIT TERMENUL ({len(overdue)}):\n"
            + ("\n".join(overdue[:10]) or "Niciuna") + "\n\n"
            f"LIVRARE IN 1-7 ZILE ({len(next7)}):\n"
            + ("\n".join(next7[:10]) or "Niciuna") + "\n\n"
            f"LIVRARE IN 8-30 ZILE ({len(next30)}):\n"
            + ("\n".join(next30[:10]) or "Niciuna") + "\n\n"
            f"PLANIFICARE: {plan_line}"
        )
        return question, context

    # ── DASHBOARD: Blocaje ────────────────────────────────────────────────────
    elif question_id == "dashboard_blocaje":
        question = "Unde sunt principalele blocaje în producție și ce le cauzează?"
        no_mat, blocked, no_bt_list = [], [], []
        if sesiune:
            for status_filter, target in [
                ("no_material", no_mat),
                ("blocked_by_rank", blocked),
                ("no_bt", no_bt_list),
            ]:
                rows = (
                    db.query(PlanificareRezultat.wo, func.count(PlanificareRezultat.id))
                    .filter(PlanificareRezultat.sesiune_id == sesiune.id)
                    .filter(PlanificareRezultat.status == status_filter)
                    .group_by(PlanificareRezultat.wo)
                    .order_by(func.count(PlanificareRezultat.id).desc())
                    .limit(8)
                    .all()
                )
                for wo_id, cnt in rows:
                    c = db.query(Comanda).filter(Comanda.cp == wo_id).first()
                    cl = (c.client or "N/A")[:20] if c else "N/A"
                    target.append(f"WO:{wo_id} ({cnt} op) | {cl}")

        context = (
            f"DATA: {today}\n"
            f"PLANIFICARE: {plan_line}\n\n"
            f"BLOCAJ FARA MATERIAL ({plan_stats.get('no_material',0)} op):\n"
            + ("\n".join(no_mat) or "Nicio operatie") + "\n\n"
            f"BLOCAJ RANK ({plan_stats.get('blocked_by_rank',0)} op - asteapta operatii precedente):\n"
            + ("\n".join(blocked) or "Nicio operatie") + "\n\n"
            f"FARA BUN DE TIPAR ({plan_stats.get('no_bt',0)} op):\n"
            + ("\n".join(no_bt_list) or "Nicio operatie") + "\n\n"
            f"FARA RESURSA: {plan_stats.get('no_resource',0)} op"
        )
        return question, context

    # ── PLANIFICARE: De ce sunt blocate ───────────────────────────────────────
    elif question_id == "plan_blocate":
        question = "De ce sunt comenzi blocate și care este cauza principală?"
        if not sesiune:
            return question, "Nicio sesiune de planificare. Ruleaza din Dashboard."

        blocked_ops = (
            db.query(PlanificareRezultat)
            .filter(PlanificareRezultat.sesiune_id == sesiune.id)
            .filter(PlanificareRezultat.status.in_(
                ["no_material", "blocked_by_rank", "no_resource", "no_bt",
                 "blocat_semifabricat", "blocat_prefabricat"]
            ))
            .order_by(PlanificareRezultat.wo)
            .limit(60)
            .all()
        )
        wo_map: dict = {}
        for r in blocked_ops:
            if r.wo not in wo_map:
                wo_map[r.wo] = {"s": set(), "ops": [], "motiv": r.motiv}
            wo_map[r.wo]["s"].add(r.status)
            wo_map[r.wo]["ops"].append(r.op)

        lines = []
        for wo_id, info in list(wo_map.items())[:18]:
            c = db.query(Comanda).filter(Comanda.cp == wo_id).first()
            cl = (c.client or "N/A")[:20] if c else "N/A"
            d = (c.data_actualizata_livrare or c.dt_livr_prod) if c else None
            causes = "+".join(sorted(info["s"]))
            motiv = (info["motiv"] or "")[:55]
            lines.append(
                f"WO:{wo_id} | {cl} | Livr:{d} | "
                f"Cauze:{causes} | Ops:{info['ops'][:4]} | {motiv}"
            )

        total_blocked = sum(v for k, v in plan_stats.items() if k != "planned")
        context = (
            f"DATA: {today}\n"
            f"PLANIFICARE: {plan_line}\n"
            f"Total blocate: {total_blocked} din {sum(plan_stats.values())} operatii\n\n"
            f"DETALIU WO BLOCATE:\n" + "\n".join(lines)
        )
        return question, context

    # ── PLANIFICARE: Aprovizionare materiale ──────────────────────────────────
    elif question_id == "plan_aprovizionare":
        question = "Ce materiale trebuie aprovizionate urgent pentru a debloca comenzile?"
        deficits = (
            db.query(
                Deficit.articol,
                func.max(Deficit.sold_actual).label("sold"),
                func.sum(Deficit.cantitate).label("rezervat"),
            )
            .group_by(Deficit.articol)
            .all()
        )
        critical, low = [], []
        for d in deficits:
            avail = (d.sold or 0) + (d.rezervat or 0)
            art = (d.articol or "N/A")[:48]
            if avail < 0:
                critical.append(
                    f"{art} | Stoc:{d.sold:.0f} | Rez:{d.rezervat:.0f} | Deficit:{avail:.0f}"
                )
            elif avail < 10:
                low.append(
                    f"{art} | Stoc:{d.sold:.0f} | Rez:{d.rezervat:.0f} | Disp:{avail:.0f}"
                )

        context = (
            f"DATA: {today}\n"
            f"Op blocate din lipsa material: {plan_stats.get('no_material',0)}\n\n"
            f"MATERIALE IN DEFICIT NEGATIV ({len(critical)} art) — aprovizionare URGENTA:\n"
            + ("\n".join(critical[:15]) or "Niciun material") + "\n\n"
            f"MATERIALE CU STOC SCAZUT 0-10 ({len(low)} art):\n"
            + ("\n".join(low[:10]) or "Niciun material")
        )
        return question, context

    # ── PLANIFICARE: Quick wins ────────────────────────────────────────────────
    elif question_id == "plan_quickwins":
        question = "Care sunt operațiile care pot fi deblocate cel mai rapid (quick wins)?"
        if not sesiune:
            return question, "Nicio sesiune de planificare. Ruleaza din Dashboard."

        # 1. blocked_by_rank with near delivery — predecessor may already be done
        rank_blocked = (
            db.query(PlanificareRezultat)
            .filter(PlanificareRezultat.sesiune_id == sesiune.id)
            .filter(PlanificareRezultat.status == "blocked_by_rank")
            .limit(50)
            .all()
        )
        rank_lines = []
        for r in rank_blocked:
            c = db.query(Comanda).filter(Comanda.cp == r.wo).first()
            d = (c.data_actualizata_livrare or c.dt_livr_prod) if c else None
            delta = (d - today).days if d else 999
            motiv = (r.motiv or "")[:70]
            rank_lines.append(
                (delta, f"WO:{r.wo} OP:{r.op} CL:{r.cl} | Livr:{d} ({delta}z) | {motiv}")
            )
        rank_lines.sort(key=lambda x: x[0])

        # 2. small material deficits — cheap to fix
        deficits = (
            db.query(
                Deficit.articol,
                func.max(Deficit.sold_actual).label("sold"),
                func.sum(Deficit.cantitate).label("rezervat"),
            )
            .group_by(Deficit.articol)
            .all()
        )
        small_def = []
        for d in deficits:
            avail = (d.sold or 0) + (d.rezervat or 0)
            if -100 < avail < 0:
                small_def.append(
                    f"{(d.articol or '')[:48]} | Deficit mic: {avail:.0f} unitati"
                )

        context = (
            f"DATA: {today}\n"
            f"PLANIFICARE: {plan_line}\n\n"
            f"OPERATII BLOCATE-RANK CU LIVRARE APROPIATA (verifica daca predecesorul e finalizat):\n"
            + ("\n".join([x[1] for x in rank_lines[:12]]) or "Nicio operatie") + "\n\n"
            f"MATERIALE CU DEFICIT MIC (<100 unitati) — comanda rapida poate debloca:\n"
            + ("\n".join(small_def[:10]) or "Niciun material cu deficit mic")
        )
        return question, context

    # ── PLANIFICARE: Explică algoritmul ───────────────────────────────────────
    elif question_id == "plan_algoritm":
        question = "Explică în pași clari cum funcționează algoritmul de planificare."
        context = (
            f"DATA: {today}\n"
            f"PLANIFICARE: {ses_label}\n"
            f"REZULTATE: {plan_line}\n\n"
            f"ALGORITMUL DE PLANIFICARE — PAȘI:\n"
            f"1. SORTARE COMENZI: prioritate descrescătoare după StadiuPrepress "
            f"(06-În producție=100, 05-BT existent=50, 04-Trimis BT=40, 03-Fisiere=30, "
            f"02-Job creat=20, 01-Fără Fișiere=10), apoi dată livrare crescătoare, "
            f"apoi comenzile cu material înaintea celor fără.\n"
            f"2. EXCLUDERE: comenzile cu Status_cda=STOP sunt sărite complet.\n"
            f"3. VERIFICARE BT: operația necesită BT valid (bt1/bt2/bt3/bt4 nevid și ≠ 1911-11-11). "
            f"Fără BT → 'Fară BT'. Dacă există data_limita_bt → 'Previzionat (fără BT)' cu start la acea dată.\n"
            f"4. VERIFICARE RANK: operațiile unui WO au un număr de ordine (rank). "
            f"O operație cu rank N nu poate fi planificată dacă există o operație cu rank < N "
            f"care este încă 'open' (neînchisă și neplanificată). "
            f"Dacă predecesoarea este planificată, operația curentă începe după end-ul ei.\n"
            f"5. VERIFICARE MATERIAL: SoldActual + Aprovizionări_externe_fără_dată - Rezervări > 0. "
            f"Dacă lipsește materialul dar există o comandă de aprovizionare cu dată → 'Previzionat (fără material)', "
            f"start la data aprovizionării. Dacă materialul e produs intern de alt WO → 'Blocat Prefabricat'.\n"
            f"6. ALOCARE RESURSĂ: se caută toate resursele din CL (centru de lucru) compatibile cu operația. "
            f"Se selectează resursa cu cel mai devreme slot disponibil (load balancing). "
            f"Operațiile aceluiași WO sunt mereu secvențiale (nu paralele).\n"
            f"7. CALCUL TIMP: ore_necesare = max(0, P_Setup + P_Runtime - R_Runtime).\n"
            f"8. STATUT FINAL: planned / previzionat_bt / previzionat_material / previzionat_semifabricat / "
            f"no_material / blocat_semifabricat / blocked_by_rank / no_bt / no_resource.\n"
            f"   previzionat_semifabricat = materialul e produs de un alt WO intern PLANIFICAT → "
            f"productia poate incepe dupa ce cel intern se termina.\n"
            f"   blocat_semifabricat = materialul e produs intern dar acel WO NU e inca planificat.\n"
        )
        return question, context

    return "Întrebare necunoscută.", "Nu există context disponibil."


@app.get("/api/ai/analyze")
async def ai_analyze(
    question_id: str = Query(...),
    db: Session = Depends(get_db),
):
    """Stream an AI analysis via SSE. Each chunk: data: {"text": "..."}\n\n"""

    async def _error_stream(msg: str):
        yield f"data: {_json.dumps({'text': msg})}\n\n"
        yield "data: [DONE]\n\n"

    if not _ANTHROPIC_OK:
        return StreamingResponse(
            _error_stream("Pachetul 'anthropic' nu este instalat pe server. Rulati: pip install anthropic"),
            media_type="text/event-stream",
        )

    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        return StreamingResponse(
            _error_stream(
                "ANTHROPIC_API_KEY nu este configurat pe server.\n"
                "Adaugati-l in /etc/systemd/system/arta-grafica.service:\n"
                "Environment=ANTHROPIC_API_KEY=sk-ant-..."
            ),
            media_type="text/event-stream",
        )

    try:
        question, context = _build_ai_context(question_id, db)
    except Exception as exc:
        return StreamingResponse(
            _error_stream(f"Eroare la construirea contextului: {exc}"),
            media_type="text/event-stream",
        )

    async def _stream():
        try:
            client = _anthropic.AsyncAnthropic(api_key=api_key)
            async with client.messages.stream(
                model="claude-opus-4-6",
                max_tokens=1200,
                thinking={"type": "adaptive"},
                system=_AI_SYSTEM_PROMPT,
                messages=[{
                    "role": "user",
                    "content": f"CONTEXT:\n{context}\n\nÎNTREBARE: {question}",
                }],
            ) as stream:
                async for text in stream.text_stream:
                    yield f"data: {_json.dumps({'text': text})}\n\n"
            yield "data: [DONE]\n\n"
        except Exception as exc:
            yield f"data: {_json.dumps({'text': f'[Eroare API: {exc}]'})}\n\n"
            yield "data: [DONE]\n\n"

    return StreamingResponse(
        _stream(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


# ── Serve React frontend (production build) ───────────────────────────────────
_DIST = os.path.join(os.path.dirname(__file__), "..", "frontend", "dist")
if os.path.isdir(_DIST):
    app.mount("/assets", StaticFiles(directory=os.path.join(_DIST, "assets")), name="assets")

    _DIST_REAL = os.path.realpath(_DIST)

    @app.get("/{full_path:path}", include_in_schema=False)
    def serve_spa(full_path: str):
        """Catch-all: return index.html for SPA routing. Guards against path traversal."""
        _index = os.path.join(_DIST_REAL, "index.html")
        candidate = os.path.realpath(os.path.join(_DIST_REAL, full_path))
        # Block any path that escapes the dist directory
        if not candidate.startswith(_DIST_REAL + os.sep) and candidate != _DIST_REAL:
            return FileResponse(
                _index,
                headers={"Cache-Control": "no-cache, no-store, must-revalidate"},
            )
        if os.path.isfile(candidate):
            return FileResponse(candidate)
        return FileResponse(
            _index,
            headers={"Cache-Control": "no-cache, no-store, must-revalidate"},
        )


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=False)
